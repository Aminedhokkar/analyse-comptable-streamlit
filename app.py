import re
import io
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Comptaflow — Analyse post-migration",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════════════════════════════
# DESIGN SYSTEM — CSS global
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Sora:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">

<style>
/* ── Tokens ── */
:root {
  --ink:       #0D1117;
  --ink-2:     #1C2333;
  --ink-3:     #252D3D;
  --surface:   #F7F8FB;
  --card:      #FFFFFF;
  --border:    #E4E8F0;
  --border-2:  #CDD3DF;
  --accent:    #2563EB;
  --accent-lt: #EFF4FF;
  --accent-2:  #0EA5E9;
  --gold:      #F59E0B;
  --gold-lt:   #FFFBEB;
  --red:       #EF4444;
  --red-lt:    #FEF2F2;
  --green:     #10B981;
  --green-lt:  #ECFDF5;
  --muted:     #6B7A99;
  --muted-2:   #9AA3B8;
  --font:      'Sora', sans-serif;
  --mono:      'JetBrains Mono', monospace;
  --r-sm:      6px;
  --r-md:      10px;
  --r-lg:      16px;
  --r-xl:      20px;
  --shadow-sm: 0 1px 3px rgba(13,17,23,0.06), 0 1px 2px rgba(13,17,23,0.04);
  --shadow-md: 0 4px 16px rgba(13,17,23,0.08), 0 2px 6px rgba(13,17,23,0.04);
  --shadow-lg: 0 8px 32px rgba(13,17,23,0.12), 0 4px 12px rgba(13,17,23,0.06);
}

/* ── Reset & base ── */
*, *::before, *::after { box-sizing: border-box; }
html, body, [data-testid="stAppViewContainer"] {
  font-family: var(--font) !important;
  background: var(--surface) !important;
  color: var(--ink) !important;
}
[data-testid="stHeader"] { display: none !important; }
[data-testid="block-container"] {
  padding: 2rem 2.5rem 3rem !important;
  max-width: 1400px !important;
}

/* ── Sidebar ── */
[data-testid="stSidebar"] {
  background: var(--ink) !important;
  border-right: 1px solid var(--ink-3) !important;
  min-width: 240px !important;
  max-width: 240px !important;
}
[data-testid="stSidebar"] > div:first-child {
  padding: 0 !important;
}
[data-testid="stSidebarContent"] {
  padding: 0 !important;
}
section[data-testid="stSidebar"] .stButton > button {
  width: 100% !important;
  text-align: left !important;
  background: transparent !important;
  border: none !important;
  border-radius: var(--r-sm) !important;
  color: #8892A4 !important;
  font-size: 13px !important;
  font-weight: 500 !important;
  font-family: var(--font) !important;
  padding: 9px 14px !important;
  transition: all 0.15s ease !important;
  letter-spacing: 0.01em !important;
}
section[data-testid="stSidebar"] .stButton > button:hover {
  background: rgba(255,255,255,0.06) !important;
  color: #FFFFFF !important;
}
section[data-testid="stSidebar"] .stButton > button[kind="primary"] {
  background: rgba(37,99,235,0.15) !important;
  color: #FFFFFF !important;
  border-left: 2px solid var(--accent) !important;
  border-radius: 0 var(--r-sm) var(--r-sm) 0 !important;
}

/* ── Typographie ── */
h1 { font-family: var(--font) !important; font-size: 1.6rem !important; font-weight: 700 !important; color: var(--ink) !important; letter-spacing: -0.02em !important; margin-bottom: 0.2rem !important; }
h2 { font-family: var(--font) !important; font-size: 1.15rem !important; font-weight: 600 !important; color: var(--ink) !important; }
h3 { font-family: var(--font) !important; font-size: 1rem !important; font-weight: 600 !important; color: var(--ink) !important; }
p, li { font-size: 13.5px !important; line-height: 1.6 !important; }

/* ── Metric cards ── */
[data-testid="metric-container"] {
  background: var(--card) !important;
  border: 1px solid var(--border) !important;
  border-radius: var(--r-lg) !important;
  padding: 18px 20px !important;
  box-shadow: var(--shadow-sm) !important;
  transition: box-shadow 0.2s ease !important;
}
[data-testid="metric-container"]:hover {
  box-shadow: var(--shadow-md) !important;
}
[data-testid="stMetricLabel"] {
  font-family: var(--font) !important;
  font-size: 10.5px !important;
  font-weight: 600 !important;
  text-transform: uppercase !important;
  letter-spacing: 0.08em !important;
  color: var(--muted) !important;
}
[data-testid="stMetricValue"] {
  font-family: var(--font) !important;
  font-size: 26px !important;
  font-weight: 700 !important;
  color: var(--ink) !important;
  letter-spacing: -0.02em !important;
}
[data-testid="stMetricDelta"] {
  font-size: 11.5px !important;
  font-weight: 500 !important;
}

/* ── File uploader ── */
[data-testid="stFileUploader"] {
  background: var(--card) !important;
  border-radius: var(--r-lg) !important;
}
[data-testid="stFileUploader"] > section {
  background: var(--surface) !important;
  border: 1.5px dashed var(--border-2) !important;
  border-radius: var(--r-md) !important;
  transition: all 0.2s ease !important;
}
[data-testid="stFileUploader"] > section:hover {
  border-color: var(--accent) !important;
  background: var(--accent-lt) !important;
}
[data-testid="stFileUploader"] > section p {
  color: var(--muted) !important;
  font-size: 12.5px !important;
}

/* ── Tabs ── */
[data-testid="stTabs"] [role="tablist"] {
  gap: 0 !important;
  border-bottom: 1.5px solid var(--border) !important;
  padding-bottom: 0 !important;
  background: transparent !important;
}
[data-baseweb="tab"] {
  font-family: var(--font) !important;
  font-size: 12.5px !important;
  font-weight: 500 !important;
  padding: 10px 20px !important;
  border-radius: var(--r-sm) var(--r-sm) 0 0 !important;
  color: var(--muted) !important;
  transition: all 0.15s ease !important;
  letter-spacing: 0.01em !important;
}
[data-baseweb="tab"][aria-selected="true"] {
  background: var(--card) !important;
  color: var(--accent) !important;
  border-bottom: 2px solid var(--accent) !important;
  font-weight: 600 !important;
}
[data-baseweb="tab"]:hover:not([aria-selected="true"]) {
  background: rgba(37,99,235,0.04) !important;
  color: var(--ink) !important;
}
[data-testid="stTabsContent"] {
  background: var(--card) !important;
  border: 1px solid var(--border) !important;
  border-top: none !important;
  border-radius: 0 0 var(--r-lg) var(--r-lg) !important;
  padding: 20px !important;
  box-shadow: var(--shadow-sm) !important;
}

/* ── Tables ── */
[data-testid="stDataFrame"] {
  border-radius: var(--r-md) !important;
  border: 1px solid var(--border) !important;
  overflow: hidden !important;
  box-shadow: var(--shadow-sm) !important;
  font-family: var(--mono) !important;
}
[data-testid="stDataFrame"] th {
  background: var(--ink) !important;
  color: white !important;
  font-family: var(--font) !important;
  font-size: 11px !important;
  font-weight: 600 !important;
  text-transform: uppercase !important;
  letter-spacing: 0.06em !important;
}

/* ── Download button ── */
[data-testid="stDownloadButton"] > button {
  background: var(--ink) !important;
  color: white !important;
  border: none !important;
  border-radius: var(--r-md) !important;
  font-family: var(--font) !important;
  font-weight: 600 !important;
  font-size: 13.5px !important;
  padding: 12px 32px !important;
  letter-spacing: 0.01em !important;
  box-shadow: var(--shadow-md) !important;
  transition: all 0.2s ease !important;
}
[data-testid="stDownloadButton"] > button:hover {
  background: var(--ink-2) !important;
  box-shadow: var(--shadow-lg) !important;
  transform: translateY(-1px) !important;
}

/* ── Text inputs ── */
[data-testid="stTextInput"] input {
  font-family: var(--font) !important;
  border-radius: var(--r-sm) !important;
  border: 1px solid var(--border-2) !important;
  font-size: 13px !important;
  padding: 8px 12px !important;
  background: var(--surface) !important;
  color: var(--ink) !important;
  transition: all 0.15s ease !important;
}
[data-testid="stTextInput"] input:focus {
  border-color: var(--accent) !important;
  background: var(--card) !important;
  box-shadow: 0 0 0 3px rgba(37,99,235,0.10) !important;
}
[data-testid="stTextInput"] label {
  font-family: var(--font) !important;
  font-size: 11.5px !important;
  font-weight: 600 !important;
  color: var(--muted) !important;
  text-transform: uppercase !important;
  letter-spacing: 0.06em !important;
}

/* ── Alerts ── */
[data-testid="stAlert"] {
  border-radius: var(--r-md) !important;
  font-family: var(--font) !important;
  font-size: 13px !important;
  border: none !important;
  box-shadow: var(--shadow-sm) !important;
}
[data-testid="stAlert"][data-baseweb="notification"][kind="positive"] {
  background: var(--green-lt) !important;
  border-left: 3px solid var(--green) !important;
}
[data-testid="stAlert"][data-baseweb="notification"][kind="warning"] {
  background: var(--gold-lt) !important;
  border-left: 3px solid var(--gold) !important;
}
[data-testid="stAlert"][data-baseweb="notification"][kind="error"] {
  background: var(--red-lt) !important;
  border-left: 3px solid var(--red) !important;
}

/* ── Toggle ── */
[data-testid="stCheckbox"] label {
  font-family: var(--font) !important;
  font-size: 12.5px !important;
  font-weight: 500 !important;
  color: var(--muted) !important;
}

/* ── Selectbox ── */
[data-testid="stSelectbox"] > div > div {
  font-family: var(--font) !important;
  border-radius: var(--r-sm) !important;
  border: 1px solid var(--border-2) !important;
  font-size: 13px !important;
  background: var(--surface) !important;
}

/* ── Expander ── */
[data-testid="stExpander"] {
  border: 1px solid var(--border) !important;
  border-radius: var(--r-md) !important;
  background: var(--card) !important;
  margin-bottom: 6px !important;
  box-shadow: var(--shadow-sm) !important;
  transition: box-shadow 0.15s ease !important;
}
[data-testid="stExpander"]:hover {
  box-shadow: var(--shadow-md) !important;
}
[data-testid="stExpander"] summary {
  font-family: var(--font) !important;
  font-size: 13px !important;
  font-weight: 600 !important;
  padding: 12px 16px !important;
  color: var(--ink) !important;
}

/* ── Caption ── */
[data-testid="stCaptionContainer"] {
  font-family: var(--mono) !important;
  font-size: 11px !important;
  color: var(--muted-2) !important;
  letter-spacing: 0.03em !important;
}

/* ── Spinner ── */
[data-testid="stSpinner"] p {
  font-family: var(--font) !important;
  color: var(--accent) !important;
  font-size: 13px !important;
}

/* ── Divider ── */
hr {
  border: none !important;
  border-top: 1px solid var(--border) !important;
  margin: 2rem 0 !important;
}

/* ── Scrollbar ── */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: var(--border-2); border-radius: 8px; }
::-webkit-scrollbar-thumb:hover { background: var(--muted-2); }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR NAVIGATION
# ══════════════════════════════════════════════════════════════════════════════
if "menu" not in st.session_state:
    st.session_state.menu = "🏠 Accueil"

MENUS = [
    ("🏠 Accueil",              "🏠 Accueil"),
    ("📒 Grand Livre",          "📒 Grand Livre"),
    ("⚖️ Balance Auxiliaire",  "⚖️ Balance Auxiliaire"),
    ("📈 Balance Générale",     "📈 Balance Générale"),
    ("📗 Grand Livre Détaillé", "📗 Grand Livre Détaillé"),
]

with st.sidebar:
    # Logo / Brand
    st.markdown("""
    <div style="padding:28px 20px 24px;">
        <div style="display:flex;align-items:center;gap:10px;margin-bottom:4px;">
            <div style="width:30px;height:30px;background:linear-gradient(135deg,#2563EB,#0EA5E9);
                        border-radius:8px;display:flex;align-items:center;justify-content:center;
                        font-size:14px;">◈</div>
            <span style="font-family:'Sora',sans-serif;font-size:15px;font-weight:700;
                         color:#FFFFFF;letter-spacing:-0.01em;">Comptaflow</span>
        </div>
        <div style="font-family:'Sora',sans-serif;font-size:10.5px;color:#4B5563;
                    text-transform:uppercase;letter-spacing:0.08em;padding-left:2px;">
            Analyse post-migration
        </div>
    </div>

    <div style="height:1px;background:rgba(255,255,255,0.06);margin:0 16px 16px;"></div>

    <div style="padding:0 12px 8px;">
        <div style="font-family:'Sora',sans-serif;font-size:9.5px;font-weight:600;
                    color:#374151;text-transform:uppercase;letter-spacing:0.1em;
                    padding:0 6px;margin-bottom:6px;">Modules</div>
    </div>
    """, unsafe_allow_html=True)

    LABELS_SIDEBAR = [
        ("🏠", "Accueil",              "🏠 Accueil"),
        ("📒", "Grand Livre",          "📒 Grand Livre"),
        ("⚖️", "Balance Auxiliaire",  "⚖️ Balance Auxiliaire"),
        ("📈", "Balance Générale",     "📈 Balance Générale"),
        ("📗", "GL Détaillé",          "📗 Grand Livre Détaillé"),
    ]
    for icon, label, key in LABELS_SIDEBAR:
        is_active = st.session_state.menu == key
        if st.button(
            f"{icon}  {label}",
            key=f"nav_{key}",
            use_container_width=True,
            type="primary" if is_active else "secondary",
        ):
            st.session_state.menu = key
            st.rerun()

    # Footer sidebar
    st.markdown("""
    <div style="position:fixed;bottom:24px;left:0;width:240px;padding:0 20px;">
        <div style="height:1px;background:rgba(255,255,255,0.06);margin-bottom:14px;"></div>
        <div style="font-family:'Sora',sans-serif;font-size:10.5px;color:#374151;">
            ◈ Comptaflow v2.0<br>
            <span style="color:#2D3748;">Comparaison · Export Excel</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

menu = st.session_state.menu


# ══════════════════════════════════════════════════════════════════════════════
# UTILITAIRES COMMUNS
# ══════════════════════════════════════════════════════════════════════════════
def _make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_sheet(ws, df, title, tab_color):
    brd = _make_border()
    ws.title = title[:31]
    ws.sheet_properties.tabColor = tab_color
    hfill = PatternFill("solid", fgColor="0D1117")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    for ci, h in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font = hfill, hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor="F7F8FB" if ri % 2 == 0 else "FFFFFF")
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = brd
            c.fill = fill
            if isinstance(val, float):
                c.number_format = "#,##0.000"
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="left")
    for ci, col in enumerate(df.columns, 1):
        w = max(len(str(col)), df[col].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, 40)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def _to_float(s: str) -> float:
    s = str(s).strip().replace("\u00a0", "").replace(" ", "")
    if not s or s == "-":
        return 0.0
    if "." in s and "," in s:
        s = s.replace(",", "")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _highlight_ecarts(df):
    styles = pd.DataFrame("", index=df.index, columns=df.columns)
    for col in [c for c in df.columns if c.startswith("Ecart_")]:
        styles[col] = df[col].apply(
            lambda v: "background-color:#FEE2E2;color:#991B1B;font-weight:600"
            if isinstance(v, float) and v < -0.001
            else ("background-color:#D1FAE5;color:#065F46;font-weight:600"
                  if isinstance(v, float) and v > 0.001 else "")
        )
    return styles


def _excel_color_ecarts(ws, comp, red_f, green_f):
    ecart_idxs = [i + 1 for i, c in enumerate(comp.columns) if c.startswith("Ecart_")]
    for ri in range(2, len(comp) + 2):
        for ci in ecart_idxs:
            c = ws.cell(row=ri, column=ci)
            if c.value and abs(c.value) > 0.001:
                c.fill = red_f if c.value < 0 else green_f


def _excel_missing_sheet(wb, missing, cols_data, title, tab_color, la, lb, group_col, group_label_col):
    brd = _make_border()
    ws = wb.create_sheet()
    ws.title = title[:31]
    ws.sheet_properties.tabColor = tab_color
    hfill    = PatternFill("solid", fgColor="0D1117")
    hfont    = Font(color="FFFFFF", bold=True, size=11)
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    blu_fill = PatternFill("solid", fgColor="DBEAFE")

    ws.cell(row=1, column=1, value="Entité / Nom").fill = hfill
    ws.cell(row=1, column=1).font = hfont
    ws.cell(row=1, column=1).border = brd
    for ci, h in enumerate(cols_data, 2):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font = hfill, hfont
        c.alignment = Alignment(horizontal="center")
        c.border = brd
    ws.row_dimensions[1].height = 20

    sup_fill = PatternFill("solid", fgColor="EFF4FF")
    sup_font = Font(bold=True, size=11)
    cur = 2
    absent_idx = cols_data.index("Absent dans") + 2 if "Absent dans" in cols_data else None

    for grp_code, grp in missing.groupby(group_col, sort=True):
        n_cols = len(cols_data) + 1
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=n_cols)
        lbl = grp[group_label_col].iloc[0] if group_label_col in grp.columns else ""
        sc = ws.cell(row=cur, column=1, value=f"  {grp_code}  –  {lbl}")
        sc.fill, sc.font = sup_fill, sup_font
        sc.alignment = Alignment(horizontal="left", vertical="center")
        sc.border = brd
        ws.row_dimensions[cur].height = 18
        cur += 1
        for _, drow in grp.iterrows():
            if absent_idx:
                rf = red_fill if drow.get("Absent dans", "") == lb else blu_fill
            else:
                rf = red_fill
            ws.cell(row=cur, column=1, value="").fill = rf
            for ci, col in enumerate(cols_data, 2):
                val = drow[col] if col in drow.index else ""
                c = ws.cell(row=cur, column=ci, value=val)
                c.fill, c.border = rf, brd
                if isinstance(val, float):
                    c.number_format = "#,##0.000"
                    c.alignment = Alignment(horizontal="right")
                else:
                    c.alignment = Alignment(horizontal="left")
            cur += 1
        cur += 1

    ws.column_dimensions["A"].width = 35
    for ci, col in enumerate(cols_data, 2):
        w = max(len(col), missing[col].astype(str).str.len().max() if col in missing.columns else 10)
        ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, 40)
    ws.freeze_panes = "A2"

    lr = cur + 1
    ws.cell(row=lr,   column=1, value="Légende :").font = Font(bold=True)
    ws.cell(row=lr+1, column=1, value=f"  Absent dans {lb}").fill = red_fill
    ws.cell(row=lr+2, column=1, value=f"  Absent dans {la}").fill = blu_fill
    return ws


def _page_header(icon: str, title: str, subtitle: str = "", badge: str = ""):
    badge_html = f'<span style="background:var(--accent-lt);color:var(--accent);font-size:10px;font-weight:700;padding:3px 10px;border-radius:20px;letter-spacing:0.06em;text-transform:uppercase;">{badge}</span>' if badge else ""
    sub_html = f'<p style="font-size:13px;color:var(--muted);margin:0;font-weight:400;">{subtitle}</p>' if subtitle else ""
    st.markdown(f"""
    <div style="display:flex;align-items:center;justify-content:space-between;
                margin-bottom:28px;padding-bottom:22px;
                border-bottom:1px solid var(--border);">
        <div style="display:flex;align-items:center;gap:16px;">
            <div style="width:46px;height:46px;background:linear-gradient(135deg,#1C2333,#252D3D);
                        border-radius:12px;display:flex;align-items:center;justify-content:center;
                        font-size:22px;box-shadow:var(--shadow-md);">{icon}</div>
            <div>
                <h1 style="margin:0 0 3px;font-size:1.45rem !important;">{title}</h1>
                {sub_html}
            </div>
        </div>
        {badge_html}
    </div>
    """, unsafe_allow_html=True)


def _upload_panel(key_a, key_b, key_la, key_lb, label_a="Fichier A", label_b="Fichier B"):
    st.markdown("""
    <div style="background:var(--card);border:1px solid var(--border);border-radius:var(--r-xl);
                padding:24px 28px 20px;margin-bottom:24px;box-shadow:var(--shadow-sm);">
        <div style="display:flex;align-items:center;gap:8px;margin-bottom:18px;">
            <div style="width:6px;height:6px;border-radius:50%;background:var(--accent);"></div>
            <span style="font-size:11px;font-weight:700;text-transform:uppercase;
                         letter-spacing:0.1em;color:var(--muted);">Fichiers source</span>
        </div>
    """, unsafe_allow_html=True)

    c1, spacer, c2 = st.columns([1, 0.04, 1])
    with c1:
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;">
            <div style="width:8px;height:8px;border-radius:2px;background:#2563EB;"></div>
            <span style="font-size:12px;font-weight:600;color:#2563EB;letter-spacing:0.02em;">{label_a}</span>
        </div>
        """, unsafe_allow_html=True)
        f1 = st.file_uploader(f"Charger {label_a}", type=["txt"], key=key_a, label_visibility="collapsed")
        LA = st.text_input("Étiquette A", value=label_a, key=key_la, placeholder="Ex: Migration")

    with spacer:
        st.markdown('<div style="height:100%;display:flex;align-items:center;justify-content:center;"><div style="width:1px;height:80px;background:var(--border);margin:0 auto;"></div></div>', unsafe_allow_html=True)

    with c2:
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;">
            <div style="width:8px;height:8px;border-radius:2px;background:#10B981;"></div>
            <span style="font-size:12px;font-weight:600;color:#10B981;letter-spacing:0.02em;">{label_b}</span>
        </div>
        """, unsafe_allow_html=True)
        f2 = st.file_uploader(f"Charger {label_b}", type=["txt"], key=key_b, label_visibility="collapsed")
        LB = st.text_input("Étiquette B", value=label_b, key=key_lb, placeholder="Ex: Référence")

    st.markdown("</div>", unsafe_allow_html=True)
    return f1, f2, LA, LB


def _alert_ecart(nb, label="fournisseur(s)"):
    if nb > 0:
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:12px;
                    background:#FFFBEB;border:1px solid #FDE68A;border-left:4px solid #F59E0B;
                    border-radius:var(--r-md);padding:12px 16px;font-size:13px;
                    margin-bottom:14px;color:#92400E;">
            <span style="font-size:16px;">⚠️</span>
            <span><b>{nb} {label}</b> présentent au moins un écart</span>
        </div>
        """, unsafe_allow_html=True)


def _kpi_bar(*metrics):
    """metrics = list of (label, value, delta=None)"""
    cols = st.columns(len(metrics))
    for col, (label, value, *delta) in zip(cols, metrics):
        d = delta[0] if delta else None
        col.metric(label, value, delta=d, delta_color="inverse" if d else "normal")


# ══════════════════════════════════════════════════════════════════════════════
# ACCUEIL
# ══════════════════════════════════════════════════════════════════════════════
if menu == "🏠 Accueil":
    st.markdown("""
    <div style="text-align:center;padding:40px 0 32px;">
        <div style="display:inline-flex;align-items:center;justify-content:center;
                    width:64px;height:64px;
                    background:linear-gradient(135deg,#0D1117,#252D3D);
                    border-radius:18px;font-size:28px;
                    box-shadow:0 8px 32px rgba(13,17,23,0.25);margin-bottom:18px;">◈</div>
        <h1 style="font-size:2.1rem !important;margin-bottom:8px;letter-spacing:-0.03em;">
            Comptaflow
        </h1>
        <p style="color:var(--muted);font-size:14.5px;max-width:520px;margin:0 auto;line-height:1.7;">
            Comparez deux extraits comptables, détectez les écarts et les documents manquants,
            exportez en Excel en un clic.
        </p>
    </div>
    """, unsafe_allow_html=True)

    MODULES = [
        {
            "icon": "📒",
            "title": "Grand Livre",
            "key": "📒 Grand Livre",
            "accent": "#2563EB",
            "accent_lt": "#EFF4FF",
            "desc": "Comparaison agrégée par fournisseur sur deux fichiers Grand Livre.",
            "features": ["Agrégation Débit / Crédit par fournisseur", "Détection documents manquants", "Export Excel 4 onglets"],
            "badge": "Fournisseurs",
        },
        {
            "icon": "⚖️",
            "title": "Balance Auxiliaire",
            "key": "⚖️ Balance Auxiliaire",
            "accent": "#10B981",
            "accent_lt": "#ECFDF5",
            "desc": "Balance antérieure, mouvements, solde — par fournisseur.",
            "features": ["Balance ant. · Mouvements · Solde", "Fournisseurs communs & manquants", "Export Excel 5 onglets"],
            "badge": "Auxiliaire",
        },
        {
            "icon": "📈",
            "title": "Balance Générale",
            "key": "📈 Balance Générale",
            "accent": "#F59E0B",
            "accent_lt": "#FFFBEB",
            "desc": "Débit / Crédit / Solde par compte sur toute la balance.",
            "features": ["Comparaison débit · crédit · solde", "Comptes manquants & écarts", "Export Excel 5 onglets"],
            "badge": "Générale",
        },
        {
            "icon": "📗",
            "title": "Grand Livre Détaillé",
            "key": "📗 Grand Livre Détaillé",
            "accent": "#8B5CF6",
            "accent_lt": "#F5F3FF",
            "desc": "Comparaison transaction par transaction (format pipe |Date|Réf|…).",
            "features": ["Débit · Crédit · Solde final par compte", "Références manquantes", "Export Excel 4 onglets"],
            "badge": "Détaillé",
        },
    ]

    cols = st.columns(2)
    for idx, mod in enumerate(MODULES):
        with cols[idx % 2]:
            feats_html = "".join(
                f'<div style="display:flex;align-items:center;gap:8px;margin:5px 0;">'
                f'<div style="width:4px;height:4px;border-radius:50%;background:{mod["accent"]};flex-shrink:0;"></div>'
                f'<span style="font-size:12.5px;color:#374151;">{f}</span></div>'
                for f in mod["features"]
            )
            st.markdown(f"""
            <div style="background:var(--card);border:1px solid var(--border);
                        border-radius:var(--r-xl);padding:24px 24px 20px;
                        margin-bottom:16px;box-shadow:var(--shadow-sm);
                        border-top:3px solid {mod["accent"]};
                        transition:box-shadow 0.2s ease;"
                 onmouseover="this.style.boxShadow='var(--shadow-md)'"
                 onmouseout="this.style.boxShadow='var(--shadow-sm)'">
                <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;">
                    <div style="display:flex;align-items:center;gap:10px;">
                        <span style="font-size:22px;">{mod["icon"]}</span>
                        <span style="font-size:15px;font-weight:700;color:var(--ink);">{mod["title"]}</span>
                    </div>
                    <span style="background:{mod["accent_lt"]};color:{mod["accent"]};
                                 font-size:9.5px;font-weight:700;padding:3px 9px;
                                 border-radius:20px;letter-spacing:0.07em;text-transform:uppercase;">
                        {mod["badge"]}
                    </span>
                </div>
                <p style="font-size:13px;color:var(--muted);margin:0 0 14px;line-height:1.6;">
                    {mod["desc"]}
                </p>
                <div style="background:{mod["accent_lt"]};border-radius:var(--r-sm);
                            padding:12px 14px;margin-bottom:18px;">
                    {feats_html}
                </div>
            </div>
            """, unsafe_allow_html=True)
            if st.button(f"Ouvrir {mod['title']} →", key=f"home_btn_{idx}",
                         use_container_width=True, type="secondary"):
                st.session_state.menu = mod["key"]
                st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 1 — GRAND LIVRE
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "📒 Grand Livre":

    _page_header("📒", "Grand Livre Fournisseurs",
                 "Comparaison agrégée Débit / Crédit par fournisseur", "Module 1")

    @st.cache_data
    def parse_grand_livre(file_bytes: bytes, label: str = "fichier") -> pd.DataFrame:
        rows = []
        current_code = current_name = ""
        col_d = col_c = col_s = col_s_end = None

        for line in file_bytes.decode("utf-8", errors="ignore").splitlines():
            line = line.replace("\r", "")
            m = re.match(r"^(F\d+)\s+(.*)", line)
            if m:
                current_code = m.group(1).strip()
                current_name = m.group(2).strip()
                col_d = col_c = col_s = col_s_end = None
                continue
            stripped = line.rstrip()
            if re.match(r'^[-\s]+$', stripped) and stripped.count('-') > 10:
                dash_groups = [(m2.start(), m2.end()) for m2 in re.finditer(r'-+', stripped)]
                if len(dash_groups) >= 3:
                    col_d     = dash_groups[-3][0]
                    col_c     = dash_groups[-2][0]
                    col_s     = dash_groups[-1][0]
                    col_s_end = dash_groups[-1][1]
                continue
            if not re.match(r"^\d{2}/\d{2}/\d{2}", line):
                continue
            if re.search(r"(Tot du|Cumuls au|cumuls au)", line):
                continue
            if len(line) < 50:
                continue
            if col_d is not None:
                end_d    = col_c if col_c <= len(line) else len(line)
                end_c    = col_s if col_s <= len(line) else len(line)
                end_s    = col_s_end if col_s_end and col_s_end <= len(line) else len(line)
                debit  = line[col_d:end_d].strip()
                credit = line[col_c:end_c].strip()
                solde  = line[col_s:end_s].strip()
                rest   = line[:col_d].strip()
            else:
                solde  = line[-14:].strip()
                credit = line[-28:-14].strip()
                debit  = line[-42:-28].strip()
                rest   = line[:-42].strip()
            parts = rest.split()
            if len(parts) < 4:
                continue
            rows.append({"Fournisseur": current_code, "Nom": current_name,
                         "Date": parts[0], "Document": parts[1],
                         "Type": parts[2], "Reference": parts[3],
                         "Debit": debit, "Credit": credit, "Solde": solde})

        if not rows:
            st.warning(f"⚠️ **{label}** : aucune transaction détectée.")
            return pd.DataFrame(columns=["Fournisseur","Nom","Date","Document","Type","Reference","Debit","Credit","Solde"])
        df = pd.DataFrame(rows)
        for col in ["Debit","Credit","Solde"]:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(r"\s+","",regex=True).str.replace(",",".",regex=False), errors="coerce").fillna(0)
        return df

    def gl1_compute_missing(df1, df2, la, lb):
        all_sup = pd.concat([df1[["Fournisseur","Nom"]], df2[["Fournisseur","Nom"]]]).drop_duplicates("Fournisseur").sort_values("Fournisseur")
        records = []
        for _, sup_row in all_sup.iterrows():
            code = sup_row["Fournisseur"]; nom = sup_row["Nom"]
            docs1 = set(df1.loc[df1["Fournisseur"]==code,"Document"])
            docs2 = set(df2.loc[df2["Fournisseur"]==code,"Document"])
            for doc in sorted(docs1-docs2):
                r = df1[(df1["Fournisseur"]==code)&(df1["Document"]==doc)].iloc[0]
                records.append({"Fournisseur":code,"Nom":nom,"Document":doc,"Présent dans":la,"Absent dans":lb,"Date":r["Date"],"Type":r["Type"],"Reference":r["Reference"],"Debit":r["Debit"],"Credit":r["Credit"]})
            for doc in sorted(docs2-docs1):
                r = df2[(df2["Fournisseur"]==code)&(df2["Document"]==doc)].iloc[0]
                records.append({"Fournisseur":code,"Nom":nom,"Document":doc,"Présent dans":lb,"Absent dans":la,"Date":r["Date"],"Type":r["Type"],"Reference":r["Reference"],"Debit":r["Debit"],"Credit":r["Credit"]})
        if not records: return pd.DataFrame()
        return pd.DataFrame(records).sort_values(["Fournisseur","Présent dans","Document"])

    def gl1_build_excel(df1, df2, comp, missing, la, lb):
        wb = Workbook(); red_f = PatternFill("solid",fgColor="FEE2E2"); green_f = PatternFill("solid",fgColor="D1FAE5")
        ws1 = wb.active; _style_sheet(ws1,comp,"Comparaison","1F4E79"); _excel_color_ecarts(ws1,comp,red_f,green_f)
        if not missing.empty:
            _excel_missing_sheet(wb,missing,["Document","Présent dans","Absent dans","Date","Type","Reference","Debit","Credit"],"Documents manquants","C00000",la,lb,"Fournisseur","Nom")
        ws3 = wb.create_sheet(); _style_sheet(ws3,df1,f"Détail {la}","2563EB")
        ws4 = wb.create_sheet(); _style_sheet(ws4,df2,f"Détail {lb}","10B981")
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

    f1, f2, LA, LB = _upload_panel("gl1_f1","gl1_f2","gl1_la","gl1_lb","Fichier A","Fichier B")

    if f1 and f2:
        with st.spinner("Analyse en cours…"):
            df1 = parse_grand_livre(f1.read(), LA)
            df2 = parse_grand_livre(f2.read(), LB)

        if df1.empty or df2.empty:
            st.error("❌ Impossible de parser un ou plusieurs fichiers."); st.stop()

        sa, sb = f"_{LA}", f"_{LB}"
        agg1 = df1.groupby(["Fournisseur","Nom"])[["Debit","Credit"]].sum().reset_index()
        agg2 = df2.groupby(["Fournisseur","Nom"])[["Debit","Credit"]].sum().reset_index()
        comp = pd.merge(agg1,agg2,on="Fournisseur",how="outer",suffixes=(sa,sb))
        comp["Nom"] = comp[f"Nom{sa}"].fillna(comp[f"Nom{sb}"])
        comp = comp.drop(columns=[f"Nom{sa}",f"Nom{sb}"])
        comp = comp[["Fournisseur","Nom",f"Debit{sa}",f"Credit{sa}",f"Debit{sb}",f"Credit{sb}"]].fillna(0)
        comp["Ecart_Debit"]  = comp[f"Debit{sb}"]  - comp[f"Debit{sa}"]
        comp["Ecart_Credit"] = comp[f"Credit{sb}"] - comp[f"Credit{sa}"]
        missing = gl1_compute_missing(df1, df2, LA, LB)
        nb_ecart = len(comp[(comp["Ecart_Debit"].abs()>0.001)|(comp["Ecart_Credit"].abs()>0.001)])

        _kpi_bar(
            (f"Fournisseurs {LA}", df1["Fournisseur"].nunique()),
            (f"Fournisseurs {LB}", df2["Fournisseur"].nunique()),
            (f"Lignes {LA}", len(df1)),
            (f"Lignes {LB}", len(df2)),
            ("Docs manquants", len(missing) if not missing.empty else 0, "⚠️" if not missing.empty else None),
        )
        st.markdown("<div style='margin-top:4px;'></div>", unsafe_allow_html=True)

        tab1, tab2, tab3, tab4 = st.tabs(["📊 Comparaison agrégée","🔍 Documents manquants",f"📄 Détail {LA}",f"📄 Détail {LB}"])
        fmt = {c:"{:,.3f}" for c in comp.columns if comp[c].dtype==float}

        with tab1:
            col_l, col_r = st.columns([5,1])
            with col_r:
                only_ecart = st.toggle("Écarts uniquement", value=False, key="gl1_toggle")
            display = comp.copy()
            if only_ecart:
                display = display[(display["Ecart_Debit"].abs()>0.001)|(display["Ecart_Credit"].abs()>0.001)]
                st.caption(f"{len(display)} fournisseur(s) avec écart")
            _alert_ecart(nb_ecart)
            st.dataframe(display.style.format(fmt).apply(_highlight_ecarts,axis=None), use_container_width=True)

        with tab2:
            if missing.empty:
                st.success("✅ Aucun document manquant.")
            else:
                st.info(f"**{len(missing)} document(s) manquant(s)** · {missing['Fournisseur'].nunique()} fournisseur(s) concerné(s)")
                for sup_code, grp in missing.groupby("Fournisseur",sort=True):
                    nom = grp["Nom"].iloc[0]
                    only_a = grp[grp["Absent dans"]==LB]; only_b = grp[grp["Absent dans"]==LA]
                    lbl = (f"**{sup_code}** — {nom}  "
                           +(f"🔴 {len(only_a)} dans {LB}  " if not only_a.empty else "")
                           +(f"🔵 {len(only_b)} dans {LA}" if not only_b.empty else ""))
                    with st.expander(lbl):
                        cols_d=["Document","Date","Type","Reference","Debit","Credit"]
                        if not only_a.empty:
                            st.markdown(f"🔴 **Présents dans {LA} — Absents dans {LB}**")
                            st.dataframe(only_a[cols_d].reset_index(drop=True),use_container_width=True)
                        if not only_b.empty:
                            st.markdown(f"🔵 **Présents dans {LB} — Absents dans {LA}**")
                            st.dataframe(only_b[cols_d].reset_index(drop=True),use_container_width=True)

        with tab3:
            st.caption(f"{len(df1)} transactions")
            st.dataframe(df1.style.format({c:"{:,.3f}" for c in df1.columns if df1[c].dtype==float}),use_container_width=True)

        with tab4:
            st.caption(f"{len(df2)} transactions")
            st.dataframe(df2.style.format({c:"{:,.3f}" for c in df2.columns if df2[c].dtype==float}),use_container_width=True)

        st.divider()
        st.download_button("📥  Télécharger rapport Excel — 4 onglets",
                           data=gl1_build_excel(df1,df2,comp,missing,LA,LB),
                           file_name="grand_livre_comparaison.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 2 — BALANCE AUXILIAIRE
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "⚖️ Balance Auxiliaire":

    _page_header("⚖️", "Balance Auxiliaire Fournisseurs",
                 "Format 2 lignes par fournisseur — Balance ant. · Mouvements · Solde", "Module 2")

    @st.cache_data
    def parse_balance(file_bytes: bytes, label: str = "fichier") -> pd.DataFrame:
        def extract_trailing_number(seg: str) -> float:
            m = re.search(r'\s{2,}([\d,.]+)\s*$', seg)
            if m: return _to_float(m.group(1))
            seg_stripped = seg.strip()
            if re.match(r'^[\d,.]+$', seg_stripped): return _to_float(seg_stripped)
            return 0.0
        def clean_name(raw: str) -> str:
            return re.sub(r'\s+[\d,.]+\s*$',"",raw).strip()

        rows=[]; lines=[l.replace("\r","") for l in file_bytes.decode("utf-8",errors="ignore").splitlines()]
        i=0
        while i<len(lines):
            line=lines[i]
            if re.match(r"^[Ff]\d+",line):
                fline=line; j=i+1
                while j<len(lines) and not lines[j].strip(): j+=1
                nline=lines[j] if j<len(lines) else ""
                fparts=fline.split("|"); nparts=nline.split("|")
                code_m=re.match(r"^([Ff]\d+)",fparts[0]); code=code_m.group(1).upper() if code_m else ""
                name=clean_name(nparts[0]) if nparts else ""
                rows.append({"Fournisseur":code,"Nom":name,
                             "BalAnt_Debit":extract_trailing_number(fparts[0]),
                             "BalAnt_Credit":extract_trailing_number(nparts[0]) if nparts else 0.0,
                             "Mvt_Debit":_to_float(fparts[1]) if len(fparts)>1 else 0.0,
                             "Mvt_Credit":_to_float(nparts[1]) if len(nparts)>1 else 0.0,
                             "Bal_Debit":_to_float(fparts[2]) if len(fparts)>2 else 0.0,
                             "Bal_Credit":_to_float(nparts[2]) if len(nparts)>2 else 0.0,
                             "Solde_Debit":_to_float(fparts[3]) if len(fparts)>3 else 0.0,
                             "Solde_Credit":_to_float(nparts[3]) if len(nparts)>3 else 0.0})
                i=j+1
            else: i+=1
        if not rows:
            st.warning(f"⚠️ **{label}** : aucun fournisseur détecté.")
            return pd.DataFrame(columns=["Fournisseur","Nom","BalAnt_Debit","BalAnt_Credit","Mvt_Debit","Mvt_Credit","Bal_Debit","Bal_Credit","Solde_Debit","Solde_Credit"])
        return pd.DataFrame(rows)

    def ba_compute_common(df1,df2,la,lb):
        codes=set(df1["Fournisseur"])&set(df2["Fournisseur"]); sa,sb=f"_{la}",f"_{lb}"
        merged=pd.merge(df1[df1["Fournisseur"].isin(codes)],df2[df2["Fournisseur"].isin(codes)],on="Fournisseur",suffixes=(sa,sb))
        merged["Nom"]=merged[f"Nom{sa}"].fillna(merged[f"Nom{sb}"]); merged=merged.drop(columns=[f"Nom{sa}",f"Nom{sb}"])
        cols=["Fournisseur","Nom"]+[c for c in merged.columns if c not in ("Fournisseur","Nom")]
        return merged[cols].sort_values("Fournisseur").reset_index(drop=True)

    def ba_compute_missing(df1,df2,la,lb):
        codes1,codes2=set(df1["Fournisseur"]),set(df2["Fournisseur"]); records=[]
        for code in sorted(codes1-codes2):
            r=df1[df1["Fournisseur"]==code].iloc[0]
            records.append({"Fournisseur":code,"Nom":r["Nom"],"Présent dans":la,"Absent dans":lb,**{k:r[k] for k in ["BalAnt_Debit","BalAnt_Credit","Mvt_Debit","Mvt_Credit","Bal_Debit","Bal_Credit","Solde_Debit","Solde_Credit"]}})
        for code in sorted(codes2-codes1):
            r=df2[df2["Fournisseur"]==code].iloc[0]
            records.append({"Fournisseur":code,"Nom":r["Nom"],"Présent dans":lb,"Absent dans":la,**{k:r[k] for k in ["BalAnt_Debit","BalAnt_Credit","Mvt_Debit","Mvt_Credit","Bal_Debit","Bal_Credit","Solde_Debit","Solde_Credit"]}})
        if not records: return pd.DataFrame()
        return pd.DataFrame(records).sort_values(["Absent dans","Fournisseur"])

    def ba_build_excel(df1,df2,comp,missing,common,la,lb):
        wb=Workbook(); red_f=PatternFill("solid",fgColor="FEE2E2"); green_f=PatternFill("solid",fgColor="D1FAE5")
        ws1=wb.active; _style_sheet(ws1,comp,"Comparaison","1F4E79"); _excel_color_ecarts(ws1,comp,red_f,green_f)
        ws_com=wb.create_sheet(); _style_sheet(ws_com,common,"Fournisseurs communs","10B981")
        if not missing.empty:
            _excel_missing_sheet(wb,missing,["Présent dans","Absent dans","BalAnt_Debit","BalAnt_Credit","Mvt_Debit","Mvt_Credit","Bal_Debit","Bal_Credit","Solde_Debit","Solde_Credit"],"Fournisseurs manquants","C00000",la,lb,"Fournisseur","Nom")
        ws4=wb.create_sheet(); _style_sheet(ws4,df1,f"Données {la}","2563EB")
        ws5=wb.create_sheet(); _style_sheet(ws5,df2,f"Données {lb}","10B981")
        buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

    f1, f2, LA, LB = _upload_panel("ba_f1","ba_f2","ba_la","ba_lb","Fichier A","Fichier B")

    if f1 and f2:
        with st.spinner("Analyse en cours…"):
            df1=parse_balance(f1.read(),LA); df2=parse_balance(f2.read(),LB)
        if df1.empty or df2.empty:
            st.error("❌ Impossible de parser."); st.stop()

        sa,sb=f"_{LA}",f"_{LB}"
        comp=pd.merge(df1,df2,on="Fournisseur",how="outer",suffixes=(sa,sb))
        comp["Nom"]=comp[f"Nom{sa}"].fillna(comp[f"Nom{sb}"]); comp=comp.drop(columns=[f"Nom{sa}",f"Nom{sb}"]).fillna(0)
        base_cols=["Fournisseur","Nom",f"BalAnt_Debit{sa}",f"BalAnt_Credit{sa}",f"BalAnt_Debit{sb}",f"BalAnt_Credit{sb}",f"Mvt_Debit{sa}",f"Mvt_Credit{sa}",f"Mvt_Debit{sb}",f"Mvt_Credit{sb}",f"Bal_Debit{sa}",f"Bal_Credit{sa}",f"Bal_Debit{sb}",f"Bal_Credit{sb}",f"Solde_Debit{sa}",f"Solde_Credit{sa}",f"Solde_Debit{sb}",f"Solde_Credit{sb}"]
        comp=comp[base_cols]
        comp["Ecart_BalAnt_Debit"]=comp[f"BalAnt_Debit{sb}"]-comp[f"BalAnt_Debit{sa}"]
        comp["Ecart_BalAnt_Credit"]=comp[f"BalAnt_Credit{sb}"]-comp[f"BalAnt_Credit{sa}"]
        comp["Ecart_Mvt_Debit"]=comp[f"Mvt_Debit{sb}"]-comp[f"Mvt_Debit{sa}"]
        comp["Ecart_Mvt_Credit"]=comp[f"Mvt_Credit{sb}"]-comp[f"Mvt_Credit{sa}"]
        comp["Ecart_Bal_Debit"]=comp[f"Bal_Debit{sb}"]-comp[f"Bal_Debit{sa}"]
        comp["Ecart_Bal_Credit"]=comp[f"Bal_Credit{sb}"]-comp[f"Bal_Credit{sa}"]
        comp["Ecart_Solde_Debit"]=comp[f"Solde_Debit{sb}"]-comp[f"Solde_Debit{sa}"]
        comp["Ecart_Solde_Credit"]=comp[f"Solde_Credit{sb}"]-comp[f"Solde_Credit{sa}"]
        missing=ba_compute_missing(df1,df2,LA,LB); common=ba_compute_common(df1,df2,LA,LB)
        nb_ecart=len(comp[(comp["Ecart_BalAnt_Debit"].abs()>0.01)|(comp["Ecart_BalAnt_Credit"].abs()>0.01)|(comp["Ecart_Mvt_Debit"].abs()>0.01)|(comp["Ecart_Mvt_Credit"].abs()>0.01)])

        _kpi_bar(
            (f"Fournisseurs {LA}", len(df1)),
            (f"Fournisseurs {LB}", len(df2)),
            ("Communs", len(set(df1["Fournisseur"])&set(df2["Fournisseur"]))),
            (f"Uniq. {LA}", len(set(df1["Fournisseur"])-set(df2["Fournisseur"]))),
            (f"Uniq. {LB}", len(set(df2["Fournisseur"])-set(df1["Fournisseur"]))),
            ("Avec écart", nb_ecart, "⚠️" if nb_ecart>0 else None),
        )
        st.markdown("<div style='margin-top:4px;'></div>", unsafe_allow_html=True)
        fmt={c:"{:,.3f}" for c in comp.columns if comp[c].dtype==float}

        tab1,tab2,tab3,tab4,tab5=st.tabs(["📊 Comparaison","🤝 Fournisseurs communs","🔍 Fournisseurs manquants",f"📄 Données {LA}",f"📄 Données {LB}"])

        with tab1:
            col_l,col_r=st.columns([5,1])
            with col_r: only_ecart=st.toggle("Écarts uniquement",value=False,key="ba_toggle")
            display=comp.copy()
            if only_ecart:
                display=display[(display["Ecart_BalAnt_Debit"].abs()>0.01)|(display["Ecart_BalAnt_Credit"].abs()>0.01)|(display["Ecart_Mvt_Debit"].abs()>0.01)|(display["Ecart_Mvt_Credit"].abs()>0.01)]
                st.caption(f"{len(display)} fournisseur(s) avec écart")
            _alert_ecart(nb_ecart)
            st.dataframe(display.style.format(fmt).apply(_highlight_ecarts,axis=None),use_container_width=True)

        with tab2:
            st.caption(f"{len(common)} fournisseurs communs aux deux fichiers")
            st.dataframe(common.style.format({c:"{:,.3f}" for c in common.columns if common[c].dtype==float}),use_container_width=True)

        with tab3:
            if missing.empty:
                st.success("✅ Tous les fournisseurs sont présents dans les deux fichiers.")
            else:
                only_a=missing[missing["Absent dans"]==LB]; only_b=missing[missing["Absent dans"]==LA]
                st.info(f"**{len(missing)} fournisseur(s)** — 🔴 {len(only_a)} absents dans {LB} · 🔵 {len(only_b)} absents dans {LA}")
                fmt_m={c:"{:,.3f}" for c in missing.columns if missing[c].dtype==float}
                if not only_a.empty:
                    st.markdown(f"#### 🔴 Présents dans {LA} — Absents dans {LB}")
                    st.dataframe(only_a.drop(columns=["Présent dans","Absent dans"]).reset_index(drop=True).style.format(fmt_m),use_container_width=True)
                if not only_b.empty:
                    st.markdown(f"#### 🔵 Présents dans {LB} — Absents dans {LA}")
                    st.dataframe(only_b.drop(columns=["Présent dans","Absent dans"]).reset_index(drop=True).style.format(fmt_m),use_container_width=True)

        with tab4:
            st.caption(f"{len(df1)} fournisseurs")
            st.dataframe(df1.style.format({c:"{:,.3f}" for c in df1.columns if df1[c].dtype==float}),use_container_width=True)

        with tab5:
            st.caption(f"{len(df2)} fournisseurs")
            st.dataframe(df2.style.format({c:"{:,.3f}" for c in df2.columns if df2[c].dtype==float}),use_container_width=True)

        st.divider()
        st.download_button("📥  Télécharger rapport Excel — 5 onglets",
                           data=ba_build_excel(df1,df2,comp,missing,common,LA,LB),
                           file_name="balance_auxiliaire_comparaison.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 3 — BALANCE GÉNÉRALE
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "📈 Balance Générale":

    _page_header("📈", "Balance Générale",
                 "3 sections : balance antérieure · mouvements · solde final", "Module 3")

    @st.cache_data
    def parse_balance_generale(file_bytes: bytes, label: str = "fichier") -> pd.DataFrame:
        lines=[l.replace("\r","") for l in file_bytes.decode("utf-8",errors="ignore").splitlines()]
        def parse_pipe_values(line):
            cells=line.split("|"); vals=[]
            for cell in cells[2:]:
                cell=cell.strip()
                vals.append(_to_float(cell) if cell else None)
            return vals
        def is_page_header(line):
            s=line.strip()
            if not s: return True
            if not s.startswith("|"): return True
            if re.match(r"^\|[-=\s]+\|",s): return True
            if re.match(r"^\|Compte\s*\|",s,re.IGNORECASE): return True
            if re.match(r"^\|Description\s*\|",s,re.IGNORECASE): return True
            return False
        def is_description_line(line,current_code=""):
            if not line.startswith("|"): return False
            if current_code and re.match(r"^\|"+re.escape(current_code)+r"\s*\|",line): return True
            if re.match(r"^\|\d{5,}",line): return False
            if re.match(r"^\|[-=\s]+\|",line): return False
            if line.count("|")<2: return False
            return True
        rows=[]; i=0
        while i<len(lines):
            line=lines[i]; m=re.match(r"^\|(\d{8})\s*\|",line)
            if m:
                code=m.group(1); vals1=parse_pipe_values(line)
                j=i+1
                while j<len(lines) and not lines[j].strip(): j+=1
                while j<len(lines) and is_page_header(lines[j]): j+=1
                if j<len(lines) and is_description_line(lines[j],code):
                    dm=re.match(r"^\|([^|]+)\|",lines[j])
                    raw_desc=dm.group(1).strip() if dm else ""
                    description="" if raw_desc==code else raw_desc
                    vals2=parse_pipe_values(lines[j]); i=j+1
                else:
                    description=""; vals2=[None]*6; i+=1
                def pick(lst,idx): return lst[idx] if lst and idx<len(lst) and lst[idx] is not None else 0.0
                rows.append({"Compte":code,"Description":description,"BalAnt_Debit":pick(vals1,0),"BalAnt_Credit":pick(vals2,1),"Mvt_Debit":pick(vals1,2),"Mvt_Credit":pick(vals2,3),"Solde_Debit":pick(vals1,4),"Solde_Credit":pick(vals2,5)})
            else: i+=1
        if not rows:
            st.warning(f"⚠️ **{label}** : aucun compte détecté.")
            return pd.DataFrame(columns=["Compte","Description","BalAnt_Debit","BalAnt_Credit","Mvt_Debit","Mvt_Credit","Solde_Debit","Solde_Credit"])
        return pd.DataFrame(rows).sort_values("Compte").reset_index(drop=True)

    def bg_compute_missing(df1,df2,la,lb):
        codes1,codes2=set(df1["Compte"]),set(df2["Compte"]); records=[]
        for code in sorted(codes1-codes2):
            r=df1[df1["Compte"]==code].iloc[0]
            records.append({"Compte":code,"Description":r["Description"],"Présent dans":la,"Absent dans":lb,**{k:r[k] for k in ["BalAnt_Debit","BalAnt_Credit","Mvt_Debit","Mvt_Credit","Solde_Debit","Solde_Credit"]}})
        for code in sorted(codes2-codes1):
            r=df2[df2["Compte"]==code].iloc[0]
            records.append({"Compte":code,"Description":r["Description"],"Présent dans":lb,"Absent dans":la,**{k:r[k] for k in ["BalAnt_Debit","BalAnt_Credit","Mvt_Debit","Mvt_Credit","Solde_Debit","Solde_Credit"]}})
        if not records: return pd.DataFrame()
        return pd.DataFrame(records).sort_values(["Absent dans","Compte"]).reset_index(drop=True)

    def bg_build_excel(df1,df2,comp,missing,la,lb):
        wb=Workbook(); red_f=PatternFill("solid",fgColor="FEE2E2"); green_f=PatternFill("solid",fgColor="D1FAE5")
        ws1=wb.active; _style_sheet(ws1,comp,"Comparaison","1F4E79"); _excel_color_ecarts(ws1,comp,red_f,green_f)
        codes_communs=set(df1["Compte"])&set(df2["Compte"]); df_commun=comp[comp["Compte"].isin(codes_communs)].copy()
        ws_com=wb.create_sheet(); _style_sheet(ws_com,df_commun,"Comptes communs","10B981"); _excel_color_ecarts(ws_com,df_commun.reset_index(drop=True),red_f,green_f)
        if not missing.empty:
            _excel_missing_sheet(wb,missing,["Présent dans","Absent dans","BalAnt_Debit","BalAnt_Credit","Mvt_Debit","Mvt_Credit","Solde_Debit","Solde_Credit"],"Comptes manquants","C00000",la,lb,"Compte","Description")
        ws4=wb.create_sheet(); _style_sheet(ws4,df1,f"Données {la}","2563EB")
        ws5=wb.create_sheet(); _style_sheet(ws5,df2,f"Données {lb}","10B981")
        buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

    f1,f2,LA,LB=_upload_panel("bg_f1","bg_f2","bg_la","bg_lb","Fichier A (Balance Gén.)","Fichier B (Balance Gén.)")

    if f1 and f2:
        with st.spinner("Analyse en cours…"):
            df1=parse_balance_generale(f1.read(),LA); df2=parse_balance_generale(f2.read(),LB)
        if df1.empty or df2.empty:
            st.error("❌ Impossible de parser."); st.stop()

        sa,sb=f"_{LA}",f"_{LB}"
        comp=pd.merge(df1,df2,on="Compte",how="outer",suffixes=(sa,sb))
        comp["Description"]=comp[f"Description{sa}"].fillna(comp[f"Description{sb}"]); comp=comp.drop(columns=[f"Description{sa}",f"Description{sb}"]).fillna(0)
        base_cols=["Compte","Description",f"BalAnt_Debit{sa}",f"BalAnt_Credit{sa}",f"BalAnt_Debit{sb}",f"BalAnt_Credit{sb}",f"Mvt_Debit{sa}",f"Mvt_Credit{sa}",f"Mvt_Debit{sb}",f"Mvt_Credit{sb}",f"Solde_Debit{sa}",f"Solde_Credit{sa}",f"Solde_Debit{sb}",f"Solde_Credit{sb}"]
        comp=comp[base_cols]
        comp["Ecart_BalAnt_Debit"]=comp[f"BalAnt_Debit{sb}"]-comp[f"BalAnt_Debit{sa}"]
        comp["Ecart_BalAnt_Credit"]=comp[f"BalAnt_Credit{sb}"]-comp[f"BalAnt_Credit{sa}"]
        comp["Ecart_Mvt_Debit"]=comp[f"Mvt_Debit{sb}"]-comp[f"Mvt_Debit{sa}"]
        comp["Ecart_Mvt_Credit"]=comp[f"Mvt_Credit{sb}"]-comp[f"Mvt_Credit{sa}"]
        comp["Ecart_Solde_Debit"]=comp[f"Solde_Debit{sb}"]-comp[f"Solde_Debit{sa}"]
        comp["Ecart_Solde_Credit"]=comp[f"Solde_Credit{sb}"]-comp[f"Solde_Credit{sa}"]
        comp=comp.sort_values("Compte").reset_index(drop=True)
        missing=bg_compute_missing(df1,df2,LA,LB); codes_communs=set(df1["Compte"])&set(df2["Compte"])
        nb_ecart=len(comp[(comp["Ecart_BalAnt_Debit"].abs()>0.001)|(comp["Ecart_BalAnt_Credit"].abs()>0.001)|(comp["Ecart_Mvt_Debit"].abs()>0.001)|(comp["Ecart_Mvt_Credit"].abs()>0.001)|(comp["Ecart_Solde_Debit"].abs()>0.001)|(comp["Ecart_Solde_Credit"].abs()>0.001)])

        _kpi_bar(
            (f"Comptes {LA}", len(df1)),
            (f"Comptes {LB}", len(df2)),
            ("Communs", len(codes_communs)),
            (f"Uniq. {LA}", len(set(df1["Compte"])-set(df2["Compte"]))),
            (f"Uniq. {LB}", len(set(df2["Compte"])-set(df1["Compte"]))),
            ("Avec écart", nb_ecart, "⚠️" if nb_ecart>0 else None),
        )
        st.markdown("<div style='margin-top:4px;'></div>", unsafe_allow_html=True)
        fmt={c:"{:,.3f}" for c in comp.columns if comp[c].dtype==float}

        tab1,tab2,tab3,tab4,tab5=st.tabs(["📊 Comparaison","🤝 Comptes communs","🔍 Comptes manquants",f"📄 Données {LA}",f"📄 Données {LB}"])

        with tab1:
            col_l,col_r=st.columns([5,1])
            with col_r: only_ecart=st.toggle("Écarts uniquement",value=False,key="bg_toggle")
            display=comp.copy()
            if only_ecart:
                display=display[(display["Ecart_BalAnt_Debit"].abs()>0.001)|(display["Ecart_BalAnt_Credit"].abs()>0.001)|(display["Ecart_Mvt_Debit"].abs()>0.001)|(display["Ecart_Mvt_Credit"].abs()>0.001)|(display["Ecart_Solde_Debit"].abs()>0.001)|(display["Ecart_Solde_Credit"].abs()>0.001)]
                st.caption(f"{len(display)} compte(s) avec écart")
            _alert_ecart(nb_ecart,"compte(s)")
            st.dataframe(display.style.format(fmt).apply(_highlight_ecarts,axis=None),use_container_width=True)

        with tab2:
            df_commun_disp=comp[comp["Compte"].isin(codes_communs)].copy()
            st.caption(f"{len(df_commun_disp)} comptes présents dans les deux fichiers")
            st.dataframe(df_commun_disp.style.format(fmt).apply(_highlight_ecarts,axis=None),use_container_width=True)

        with tab3:
            if missing.empty:
                st.success("✅ Tous les comptes sont présents.")
            else:
                only_a=missing[missing["Absent dans"]==LB]; only_b=missing[missing["Absent dans"]==LA]
                st.info(f"**{len(missing)} compte(s)** — 🔴 {len(only_a)} absents dans {LB} · 🔵 {len(only_b)} absents dans {LA}")
                fmt_m={c:"{:,.3f}" for c in missing.columns if missing[c].dtype==float}
                if not only_a.empty:
                    st.markdown(f"#### 🔴 Présents dans {LA} — Absents dans {LB}")
                    st.dataframe(only_a.drop(columns=["Présent dans","Absent dans"]).reset_index(drop=True).style.format(fmt_m),use_container_width=True)
                if not only_b.empty:
                    st.markdown(f"#### 🔵 Présents dans {LB} — Absents dans {LA}")
                    st.dataframe(only_b.drop(columns=["Présent dans","Absent dans"]).reset_index(drop=True).style.format(fmt_m),use_container_width=True)

        with tab4:
            st.caption(f"{len(df1)} comptes")
            st.dataframe(df1.style.format({c:"{:,.3f}" for c in df1.columns if df1[c].dtype==float}),use_container_width=True)

        with tab5:
            st.caption(f"{len(df2)} comptes")
            st.dataframe(df2.style.format({c:"{:,.3f}" for c in df2.columns if df2[c].dtype==float}),use_container_width=True)

        st.divider()
        st.download_button("📥  Télécharger rapport Excel — 5 onglets",
                           data=bg_build_excel(df1,df2,comp,missing,LA,LB),
                           file_name="balance_generale_comparaison.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 4 — GRAND LIVRE DÉTAILLÉ
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "📗 Grand Livre Détaillé":

    _page_header("📗", "Grand Livre Détaillé",
                 "Comparaison transaction par transaction — format pipe |Date|Réf|Type|Libellé|D|C|", "Module 4")

    @st.cache_data
    def parse_grand_livre_detail(file_bytes: bytes, label: str = "fichier") -> pd.DataFrame:
        lines=[l.replace("\r","") for l in file_bytes.decode("utf-8",errors="ignore").splitlines()]
        rows=[]; current_code=current_desc=""
        for line in lines:
            m=re.match(r"^compte\s+(\d+)\s+(.*)",line,re.IGNORECASE)
            if m: current_code=m.group(1).strip(); current_desc=m.group(2).strip(); continue
            if re.match(r"^\d{2}/\d{2}/\d{4}\|",line):
                parts=line.split("|")
                if len(parts)<6: continue
                rows.append({"Compte":current_code,"Description":current_desc,"Date":parts[0].strip(),"Reference":parts[1].strip(),"Type":parts[2].strip(),"Libelle":parts[3].strip(),"Debit":_to_float(parts[4]),"Credit":_to_float(parts[5]),"Solde":_to_float(parts[6]) if len(parts)>6 else 0.0})
        if not rows:
            st.warning(f"⚠️ **{label}** : aucune transaction détectée.")
            return pd.DataFrame(columns=["Compte","Description","Date","Reference","Type","Libelle","Debit","Credit","Solde"])
        df=pd.DataFrame(rows)
        for col in ["Debit","Credit","Solde"]: df[col]=pd.to_numeric(df[col],errors="coerce").fillna(0.0)
        return df

    def gld_compute_missing(df1,df2,la,lb):
        all_comptes=pd.concat([df1[["Compte","Description"]],df2[["Compte","Description"]]]).drop_duplicates("Compte").sort_values("Compte")
        records=[]
        for _,row in all_comptes.iterrows():
            code=row["Compte"]; desc=row["Description"]
            refs1=set(df1.loc[df1["Compte"]==code,"Reference"]); refs2=set(df2.loc[df2["Compte"]==code,"Reference"])
            for ref in sorted(refs1-refs2):
                r=df1[(df1["Compte"]==code)&(df1["Reference"]==ref)].iloc[0]
                records.append({"Compte":code,"Description":desc,"Reference":ref,"Présent dans":la,"Absent dans":lb,"Date":r["Date"],"Type":r["Type"],"Libelle":r["Libelle"],"Debit":r["Debit"],"Credit":r["Credit"]})
            for ref in sorted(refs2-refs1):
                r=df2[(df2["Compte"]==code)&(df2["Reference"]==ref)].iloc[0]
                records.append({"Compte":code,"Description":desc,"Reference":ref,"Présent dans":lb,"Absent dans":la,"Date":r["Date"],"Type":r["Type"],"Libelle":r["Libelle"],"Debit":r["Debit"],"Credit":r["Credit"]})
        if not records: return pd.DataFrame()
        return pd.DataFrame(records).sort_values(["Compte","Présent dans","Reference"]).reset_index(drop=True)

    def gld_build_excel(df1,df2,comp,missing,la,lb):
        wb=Workbook(); red_f=PatternFill("solid",fgColor="FEE2E2"); green_f=PatternFill("solid",fgColor="D1FAE5")
        ws1=wb.active; _style_sheet(ws1,comp,"Comparaison","1F4E79"); _excel_color_ecarts(ws1,comp,red_f,green_f)
        if not missing.empty:
            _excel_missing_sheet(wb,missing,["Reference","Présent dans","Absent dans","Date","Type","Libelle","Debit","Credit"],"Références manquantes","C00000",la,lb,"Compte","Description")
        ws3=wb.create_sheet(); _style_sheet(ws3,df1,f"Détail {la}","2563EB")
        ws4=wb.create_sheet(); _style_sheet(ws4,df2,f"Détail {lb}","10B981")
        buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

    f1,f2,LA,LB=_upload_panel("gld_f1","gld_f2","gld_la","gld_lb","Fichier A (Grand Livre)","Fichier B (Grand Livre)")

    if f1 and f2:
        with st.spinner("Analyse en cours…"):
            df1=parse_grand_livre_detail(f1.read(),LA); df2=parse_grand_livre_detail(f2.read(),LB)
        if df1.empty or df2.empty:
            st.error("❌ Impossible de parser."); st.stop()

        sa,sb=f"_{LA}",f"_{LB}"
        agg1=df1.groupby(["Compte","Description"])[["Debit","Credit"]].sum().reset_index()
        solde1=df1.sort_values(["Compte","Date"]).groupby("Compte")["Solde"].last().reset_index().rename(columns={"Solde":"Solde_Final"})
        agg1=pd.merge(agg1,solde1,on="Compte",how="left")
        agg2=df2.groupby(["Compte","Description"])[["Debit","Credit"]].sum().reset_index()
        solde2=df2.sort_values(["Compte","Date"]).groupby("Compte")["Solde"].last().reset_index().rename(columns={"Solde":"Solde_Final"})
        agg2=pd.merge(agg2,solde2,on="Compte",how="left")
        comp=pd.merge(agg1,agg2,on="Compte",how="outer",suffixes=(sa,sb))
        comp["Description"]=comp[f"Description{sa}"].fillna(comp[f"Description{sb}"]); comp=comp.drop(columns=[f"Description{sa}",f"Description{sb}"]).fillna(0)
        comp=comp[["Compte","Description",f"Debit{sa}",f"Credit{sa}",f"Solde_Final{sa}",f"Debit{sb}",f"Credit{sb}",f"Solde_Final{sb}"]]
        comp["Ecart_Debit"]=comp[f"Debit{sb}"]-comp[f"Debit{sa}"]
        comp["Ecart_Credit"]=comp[f"Credit{sb}"]-comp[f"Credit{sa}"]
        comp["Ecart_Solde_Final"]=comp[f"Solde_Final{sb}"]-comp[f"Solde_Final{sa}"]
        comp=comp.sort_values("Compte").reset_index(drop=True)
        missing=gld_compute_missing(df1,df2,LA,LB)
        nb_ecart=len(comp[(comp["Ecart_Debit"].abs()>0.001)|(comp["Ecart_Credit"].abs()>0.001)|(comp["Ecart_Solde_Final"].abs()>0.001)])

        _kpi_bar(
            (f"Comptes {LA}", df1["Compte"].nunique()),
            (f"Comptes {LB}", df2["Compte"].nunique()),
            (f"Lignes {LA}", len(df1)),
            (f"Lignes {LB}", len(df2)),
            ("Réf. manquantes", len(missing) if not missing.empty else 0, "⚠️" if not missing.empty else None),
            ("Comptes avec écart", nb_ecart, "⚠️" if nb_ecart>0 else None),
        )
        st.markdown("<div style='margin-top:4px;'></div>", unsafe_allow_html=True)
        fmt={c:"{:,.3f}" for c in comp.columns if comp[c].dtype==float}

        tab1,tab2,tab3,tab4=st.tabs(["📊 Comparaison agrégée","🔍 Références manquantes",f"📄 Détail {LA}",f"📄 Détail {LB}"])

        with tab1:
            col_l,col_r=st.columns([5,1])
            with col_r: only_ecart=st.toggle("Écarts uniquement",value=False,key="gld_toggle")
            display=comp.copy()
            if only_ecart:
                display=display[(display["Ecart_Debit"].abs()>0.001)|(display["Ecart_Credit"].abs()>0.001)|(display["Ecart_Solde_Final"].abs()>0.001)]
                st.caption(f"{len(display)} compte(s) avec écart")
            _alert_ecart(nb_ecart,"compte(s)")
            st.dataframe(display.style.format(fmt).apply(_highlight_ecarts,axis=None),use_container_width=True)

        with tab2:
            if missing.empty:
                st.success("✅ Aucune référence manquante.")
            else:
                st.info(f"**{len(missing)} référence(s)** sur **{missing['Compte'].nunique()} compte(s)**")
                for compte_code,grp in missing.groupby("Compte",sort=True):
                    desc=grp["Description"].iloc[0]
                    only_a=grp[grp["Absent dans"]==LB]; only_b=grp[grp["Absent dans"]==LA]
                    lbl=(f"**{compte_code}** — {desc}  "
                         +(f"🔴 {len(only_a)} absents dans {LB}  " if not only_a.empty else "")
                         +(f"🔵 {len(only_b)} absents dans {LA}" if not only_b.empty else ""))
                    with st.expander(lbl):
                        cols_d=["Reference","Date","Type","Libelle","Debit","Credit"]; fmt_m={"Debit":"{:,.3f}","Credit":"{:,.3f}"}
                        if not only_a.empty:
                            st.markdown(f"🔴 **Présents dans {LA} — Absents dans {LB}**")
                            st.dataframe(only_a[cols_d].reset_index(drop=True).style.format(fmt_m),use_container_width=True)
                        if not only_b.empty:
                            st.markdown(f"🔵 **Présents dans {LB} — Absents dans {LA}**")
                            st.dataframe(only_b[cols_d].reset_index(drop=True).style.format(fmt_m),use_container_width=True)

        with tab3:
            comptes_a=["Tous"]+sorted(df1["Compte"].unique().tolist())
            sel_a=st.selectbox(f"Filtrer par compte ({LA})",comptes_a,key="gld_sel_a")
            disp1=df1 if sel_a=="Tous" else df1[df1["Compte"]==sel_a]
            st.caption(f"{len(disp1)} ligne(s)")
            st.dataframe(disp1.style.format({c:"{:,.3f}" for c in disp1.columns if disp1[c].dtype==float}),use_container_width=True)

        with tab4:
            comptes_b=["Tous"]+sorted(df2["Compte"].unique().tolist())
            sel_b=st.selectbox(f"Filtrer par compte ({LB})",comptes_b,key="gld_sel_b")
            disp2=df2 if sel_b=="Tous" else df2[df2["Compte"]==sel_b]
            st.caption(f"{len(disp2)} ligne(s)")
            st.dataframe(disp2.style.format({c:"{:,.3f}" for c in disp2.columns if disp2[c].dtype==float}),use_container_width=True)

        st.divider()
        st.download_button("📥  Télécharger rapport Excel — 4 onglets",
                           data=gld_build_excel(df1,df2,comp,missing,LA,LB),
                           file_name="grand_livre_detail_comparaison.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
